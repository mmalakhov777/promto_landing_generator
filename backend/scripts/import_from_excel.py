"""Import landing structure from Excel file (docs/Структура Промто.xlsx).

Parses all rows, groups by category, creates categories and landings.
Idempotent — can be run multiple times without creating duplicates.
"""

import asyncio
import sys
from collections import defaultdict

sys.path.insert(0, ".")

import openpyxl
from slugify import slugify
from sqlalchemy import select

from app.core.database import async_session_factory, engine, Base
from app.models.category import Category
from app.models.landing import Landing, LandingContent, LandingSection, Locale, SectionType
from app.models.settings import SiteSettings
from app.services.slug import validate_slug
import app.models  # noqa: F401

EXCEL_PATH = "docs/Структура Промто.xlsx"


def parse_excel(path: str) -> dict:
    """Parse Excel file and return {category_name: [{slug, keyword_ru, search_volume, url}, ...]}."""
    wb = openpyxl.load_workbook(path)
    ws = wb["Лист1"]

    categories = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        url = ws.cell(row, 3).value
        if not url:
            continue

        cat_name = ws.cell(row, 1).value or ""
        keyword = ws.cell(row, 4).value or ""
        volume = ws.cell(row, 5).value or 0

        # Extract slug from URL (last path segment)
        slug = url.rstrip("/").split("/")[-1]

        categories[cat_name.strip()].append({
            "slug": slug,
            "keyword_ru": keyword.strip(),
            "search_volume": int(volume),
            "url": url,
        })

    return dict(categories)


def deduplicate_slugs(entries: list[dict]) -> list[dict]:
    """[RISK-04] Keep entry with highest search_volume for duplicate slugs."""
    seen: dict[str, dict] = {}
    duplicates = []
    for entry in entries:
        slug = entry["slug"]
        if slug in seen:
            duplicates.append(slug)
            if entry["search_volume"] > seen[slug]["search_volume"]:
                seen[slug] = entry
        else:
            seen[slug] = entry
    return list(seen.values()), duplicates


async def main():
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    data = parse_excel(EXCEL_PATH)
    print(f"Parsed {sum(len(v) for v in data.values())} rows across {len(data)} categories\n")

    stats = {"imported": 0, "skipped": 0, "conflicts": 0, "categories_created": 0}

    async with async_session_factory() as session:
        # Ensure SiteSettings exists
        if not (await session.execute(select(SiteSettings).limit(1))).scalar_one_or_none():
            session.add(SiteSettings(platform_url="https://app.promto.ai"))

        for cat_name, entries in data.items():
            # Derive category slug from URL structure
            if entries:
                url_parts = entries[0]["url"].rstrip("/").split("/")
                # Category slug is the path segment after domain
                cat_slug = url_parts[3] if len(url_parts) > 3 else slugify(cat_name)
            else:
                cat_slug = slugify(cat_name)

            # [RISK-04] Deduplicate
            unique_entries, dups = deduplicate_slugs(entries)
            for d in dups:
                print(f"  WARNING: duplicate slug '{d}' in '{cat_name}' — keeping higher volume")
                stats["conflicts"] += 1

            # Get or create category
            result = await session.execute(select(Category).where(Category.slug == cat_slug))
            category = result.scalar_one_or_none()
            if not category:
                category = Category(slug=cat_slug, name_ru=cat_name, name_en="")
                session.add(category)
                await session.flush()
                stats["categories_created"] += 1
                print(f"Category created: {cat_slug} ({cat_name})")

            # First entry is the category page itself (slug = cat_slug), skip it as a landing
            landing_entries = [e for e in unique_entries if e["slug"] != cat_slug]

            for entry in landing_entries:
                slug = entry["slug"]

                # [RISK-05, RISK-01] Validate slug
                errors = validate_slug(slug)
                if errors:
                    print(f"  SKIP {slug}: {'; '.join(errors)}")
                    stats["skipped"] += 1
                    continue

                # Check if already exists
                existing = await session.execute(
                    select(Landing).where(Landing.category_id == category.id, Landing.slug == slug)
                )
                if existing.scalar_one_or_none():
                    stats["skipped"] += 1
                    continue

                landing = Landing(
                    category_id=category.id,
                    slug=slug,
                    keyword_ru=entry["keyword_ru"],
                    search_volume=entry["search_volume"],
                )
                session.add(landing)
                await session.flush()

                for locale in Locale:
                    session.add(LandingContent(landing_id=landing.id, locale=locale))
                for st in SectionType:
                    session.add(LandingSection(landing_id=landing.id, section_type=st))

                stats["imported"] += 1

        await session.commit()

    print(f"\n{'='*40}")
    print(f"Import complete!")
    print(f"  Categories created: {stats['categories_created']}")
    print(f"  Landings imported: {stats['imported']}")
    print(f"  Skipped (existing/invalid): {stats['skipped']}")
    print(f"  Slug conflicts resolved: {stats['conflicts']}")


if __name__ == "__main__":
    asyncio.run(main())
